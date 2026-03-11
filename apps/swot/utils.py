from io import BytesIO

from django.http import HttpResponse
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def gerar_excel(analise):
    """Gera planilha Excel com a análise SWOT."""
    wb = Workbook()

    # ─── Sheet 1: SWOT ───────────────────────────────────
    ws = wb.active
    ws.title = "SWOT"

    # Estilos
    header_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    sub_font = Font(name="Arial", size=10, color="666666")
    section_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    item_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    fills = {
        "header": PatternFill(start_color="8C103C", end_color="8C103C", fill_type="solid"),
        "forcas": PatternFill(start_color="059669", end_color="059669", fill_type="solid"),
        "fraquezas": PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid"),
        "oportunidades": PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid"),
        "ameacas": PatternFill(start_color="D97706", end_color="D97706", fill_type="solid"),
        "item_forcas": PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid"),
        "item_fraquezas": PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid"),
        "item_oportunidades": PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid"),
        "item_ameacas": PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid"),
    }

    # Header
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Análise SWOT — {analise.cliente.empresa}"
    ws["A1"].font = header_font
    ws["A1"].fill = fills["header"]
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Data: {analise.data_analise.strftime('%d/%m/%Y')}  |  Consultor: {analise.consultor.get_full_name()}"
    ws["A2"].font = sub_font
    ws["A2"].alignment = Alignment(horizontal="center")

    # Colunas
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 40

    # Quadrantes
    row = 4
    quadrantes = [
        ("Forças (Strengths)", analise.forcas, "forcas"),
        ("Fraquezas (Weaknesses)", analise.fraquezas, "fraquezas"),
    ]
    quadrantes2 = [
        ("Oportunidades (Opportunities)", analise.oportunidades, "oportunidades"),
        ("Ameaças (Threats)", analise.ameacas, "ameacas"),
    ]

    for pair_idx, pair in enumerate([quadrantes, quadrantes2]):
        # Headers
        for col_idx, (title, items, key) in enumerate(pair):
            col = col_idx * 2 + 1
            cell = ws.cell(row=row, column=col, value=title)
            cell.font = section_font
            cell.fill = fills[key]
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            ws.merge_cells(
                start_row=row, start_column=col,
                end_row=row, end_column=col + 1,
            )

        # Items
        max_items = max(len(pair[0][1]), len(pair[1][1]), 1)
        for i in range(max_items):
            r = row + 1 + i
            for col_idx, (title, items, key) in enumerate(pair):
                col = col_idx * 2 + 1
                val = items[i] if i < len(items) else ""
                cell = ws.cell(row=r, column=col, value=f"  {val}" if val else "")
                cell.font = item_font
                cell.fill = fills[f"item_{key}"]
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                ws.merge_cells(
                    start_row=r, start_column=col,
                    end_row=r, end_column=col + 1,
                )

        row += max_items + 2

    # ─── Sheet 2: Mercado ────────────────────────────────
    ws2 = wb.create_sheet("Mercado")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 60

    ws2.merge_cells("A1:B1")
    ws2["A1"] = "Informações de Mercado"
    ws2["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    ws2["A1"].fill = fills["header"]
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 30

    label_font = Font(name="Arial", size=10, bold=True, color="333333")
    val_font = Font(name="Arial", size=10)

    campos = [
        ("Segmento de Mercado", analise.segmento_mercado),
        ("Público-Alvo", analise.publico_alvo),
        ("Principais Concorrentes", analise.concorrentes),
        ("Diferencial Competitivo", analise.diferencial),
        ("Observações do Mercado", analise.observacoes_mercado),
    ]
    for i, (label, val) in enumerate(campos, start=3):
        ws2.cell(row=i, column=1, value=label).font = label_font
        cell = ws2.cell(row=i, column=2, value=val or "-")
        cell.font = val_font
        cell.alignment = Alignment(wrap_text=True)

    # ─── Sheet 3: Produtos ───────────────────────────────
    ws3 = wb.create_sheet("Produtos")
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 25
    ws3.column_dimensions["C"].width = 15

    ws3.merge_cells("A1:C1")
    ws3["A1"] = "Produtos Principais"
    ws3["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    ws3["A1"].fill = fills["header"]
    ws3["A1"].alignment = Alignment(horizontal="center")
    ws3.row_dimensions[1].height = 30

    for col_idx, header in enumerate(["Produto", "Categoria", "Participação (%)"], 1):
        cell = ws3.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.border = thin_border
        cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    for i, prod in enumerate(analise.produtos_principais, start=4):
        ws3.cell(row=i, column=1, value=prod.get("nome", "")).border = thin_border
        ws3.cell(row=i, column=2, value=prod.get("categoria", "")).border = thin_border
        cell = ws3.cell(row=i, column=3, value=prod.get("participacao", 0))
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"SWOT_{analise.cliente.empresa}_{analise.data_analise.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def gerar_pdf(analise):
    """Gera PDF da análise SWOT usando WeasyPrint."""
    import weasyprint

    html_string = render_to_string("swot/_report_pdf.html", {"analise": analise})
    pdf = weasyprint.HTML(string=html_string).write_pdf()

    filename = f"SWOT_{analise.cliente.empresa}_{analise.data_analise.strftime('%Y%m%d')}.pdf"
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
